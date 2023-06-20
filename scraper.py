import requests
import json
from bs4 import BeautifulSoup
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
page = 1
while page < 44:
    headers = {
        'authority': 'yandex.com',
        'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="96", "Google Chrome";v="96"',
        'device-memory': '8',
        'rtt': '200',
        'sec-ch-ua-mobile': '?0',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36',
        'viewport-width': '534',
        'accept': 'text/javascript, application/javascript, application/ecmascript, application/x-ecmascript, */*; q=0.01',
        'x-requested-with': 'XMLHttpRequest',
        'dpr': '2',
        'downlink': '10',
        'ect': '4g',
        'sec-ch-ua-platform': '"macOS"',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-mode': 'cors',
        'sec-fetch-dest': 'empty',
        'referer': 'https://yandex.com/images/search?text=persian%20miniatures%20painting',
        'accept-language': 'en-GB,en-US;q=0.9,en;q=0.8',
        'cookie': 'mda=0; yandex_gid=10614; yandexuid=7514593221640889152; yuidss=7514593221640889152; is_gdpr=0; is_gdpr_b=CIayFBDNWSgC; i=lybIvo6Itljy15THEYGO9PTLMHnYipdGO0BqIE6REaB8I79KTJjqysrt6/Mw1kkY1/185jXulG8S8OB6ZPUdsXsEPrM=; gdpr=0; _ym_uid=1640889153535659386; _ym_isad=2; ys=wprid.1640985424730352-5083407508040537587-man1-6175-man-l7-balancer-8080-BAL-8865; _ym_d=1641043302; cycada=9PMqcCnaKYeP6/TwDcFBnl8sQQNWvgXmKPLoTWzOn+k=; yp=1643481152.ygu.1#1641493957.szm.2:1440x900:937x821#1641071817.ln_tp.01; _yasc=7JlER9dg/48VVVbbx/O3heL2pyciUYTqWUZ/EPNF1L5tGiauADUvKMye',
    }

    params = (
        ('callback', 'jQuery214020010571256978538_1641043319144'),
        ('format', 'json'),
        ('request',
         '{"blocks":[{"block":"extra-content","params":{},"version":2},{"block":"serp-controller","params":{},"version":2},{"block":"serp-list_infinite_yes","params":{"initialPageNum":0},"version":2},{"block":"more_direction_next","params":{},"version":2},{"block":"gallery__items:ajax","params":{},"version":2}],"metadata":{"bundles":{"lb":"}:pKn=ctzP?b*G$lMpB4"},"assets":{"las":"justifier-vertical2=1;justifier-height=1;thumb-underlay=1;justifier-setheight=1;fitimages-height=1;justifier-fitincuts=1;react-with-dom=1;879.0=1;470e34.0=1;1191.0=1;03bb60.0=1"},"version":"0x1e136df77b3","extraContent":{"names":["i-react-ajax-adapter"]}}}'),
        ('yu', '7514593221640889152'),
        ('p', page),
        ('text', 'persian miniatures painting'),
        ('rpt', 'image'),
        ('uinfo', 'sw-1440-sh-900-ww-534-wh-821-pd-2-wp-16x10_2560x1600'),
        ('serpid', 'rUR9PmHnpLNv2QR7Mi5JQw'),
        ('serpListType', 'vertical'),
        ('thumbSnippet', '0'),
    )

    response = requests.get('https://yandex.com/images/search', headers=headers, params=params)

    # NB. Original query string below. It seems impossible to parse and
    # reproduce query strings 100% accurately so the one below is given
    # in case the reproduced version is not "correct".
    # response = requests.get('https://yandex.com/images/search?callback=jQuery214020010571256978538_1641043319144&format=json&request=%7B%22blocks%22%3A%5B%7B%22block%22%3A%22extra-content%22%2C%22params%22%3A%7B%7D%2C%22version%22%3A2%7D%2C%7B%22block%22%3A%22serp-controller%22%2C%22params%22%3A%7B%7D%2C%22version%22%3A2%7D%2C%7B%22block%22%3A%22serp-list_infinite_yes%22%2C%22params%22%3A%7B%22initialPageNum%22%3A0%7D%2C%22version%22%3A2%7D%2C%7B%22block%22%3A%22more_direction_next%22%2C%22params%22%3A%7B%7D%2C%22version%22%3A2%7D%2C%7B%22block%22%3A%22gallery__items%3Aajax%22%2C%22params%22%3A%7B%7D%2C%22version%22%3A2%7D%5D%2C%22metadata%22%3A%7B%22bundles%22%3A%7B%22lb%22%3A%22%7D%3ApKn%3DctzP%3Fb*G%24lMpB4%22%7D%2C%22assets%22%3A%7B%22las%22%3A%22justifier-vertical2%3D1%3Bjustifier-height%3D1%3Bthumb-underlay%3D1%3Bjustifier-setheight%3D1%3Bfitimages-height%3D1%3Bjustifier-fitincuts%3D1%3Breact-with-dom%3D1%3B879.0%3D1%3B470e34.0%3D1%3B1191.0%3D1%3B03bb60.0%3D1%22%7D%2C%22version%22%3A%220x1e136df77b3%22%2C%22extraContent%22%3A%7B%22names%22%3A%5B%22i-react-ajax-adapter%22%5D%7D%7D%7D&yu=7514593221640889152&p=19&text=persian+miniatures+painting&rpt=image&uinfo=sw-1440-sh-900-ww-534-wh-821-pd-2-wp-16x10_2560x1600&serpid=rUR9PmHnpLNv2QR7Mi5JQw&serpListType=vertical&thumbSnippet=0', headers=headers)

    print(type(response))

    # print(response.text)

    response = str(response.text)

    to_be_removed_1 = response.split("(")[0]
    to_be_removed_1 = to_be_removed_1 + "("

    response = response.replace(to_be_removed_1, '').strip()
    response = response[:-1]
    response = json.loads(response)

    # response = response.replace("")
    html_data = response['blocks'][2]['html']
    # print(f"html: {html_data}")

    soup = BeautifulSoup(html_data, 'html.parser')
    all_elems = soup.select("div[class*='serp-item serp-item_type_search']")
    for elem in all_elems:
        elem = elem['data-bem']
        data = json.loads(elem)['serp-item']['preview'][0]['url']
        print(data)
        row = [data]
        ws.append(row)
        print(f"\n")
    page += 1
wb.save('Data.xlsx')

#download images.............



wb_obj = openpyxl.load_workbook('Data.xlsx')

sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row=1, column=1)

print(cell_obj.value)

# print(html_data)
