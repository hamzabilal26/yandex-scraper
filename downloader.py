import random
from requests.exceptions import ProxyError
import openpyxl
import requests
import requests_cache

requests_cache.install_cache(cache_name='images', backend='sqlite', expire_after=120000)

ug_file = open("../chrome_useragents.txt", 'r')
all_ug = []
for each in ug_file.readlines():
    each = each.replace("\n", '').strip()
    all_ug.append(each)

proxy_file = open("proxies.txt")
all_proxies = []
for proxy in proxy_file.readlines():
    proxy = proxy.replace("\n", '')
    proxy = proxy.replace('"', '')
    proxy = proxy.replace(',', '').strip()
    all_proxies.append(proxy)


def get_random_proxy():
    this_proxy = random.choice(all_proxies)
    proxies = {
        'http': this_proxy,
        'https': this_proxy
    }
    return proxies


def get_useragent():
    return random.choice(all_ug)


wb_obj = openpyxl.load_workbook('Data 3.xlsx')

sheet_obj = wb_obj.active
total_rows = sheet_obj.max_row
print(total_rows)
for i in range(1, total_rows):
    if i < 470:
        continue
    while True:
        ug = get_useragent()
        proxies = get_random_proxy()
        print(f"Proxies: {proxies}")
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Referer': 'https://google.com/',
            'User-Agent': ug,
        }
        cell_obj = sheet_obj.cell(row=i, column=1)
        print(cell_obj.value)
        try:
            response = requests.get(str(cell_obj.value), headers=headers, proxies=proxies)
            file = open(f"Test/ottoman_image{i}.png", "wb")
            file.write(response.content)
            file.close()
        except ProxyError as e:
            print(e)
            continue
        except Exception as e:
            print(e)
            print('Invalid url')
            break
        else:
            break

